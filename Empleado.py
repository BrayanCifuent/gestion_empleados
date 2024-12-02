
from prettytable import PrettyTable
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sql import mydb


class Empleado:
    def __init__(self, id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento, id_rol):
        self.id_empleado = id_empleado
        self.nombre = nombre
        self.direccion = direccion
        self.telefono = telefono
        self.correo = correo
        self.fecha_inicio = fecha_inicio
        self.salario = salario
        self.id_tipo = id_tipo
        self.id_departamento = id_departamento
        self.id_rol= id_rol

class Empleado:
    def __init__(self, id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento, id_rol):
        self.id_empleado = id_empleado
        self.nombre = nombre
        self.direccion = direccion
        self.telefono = telefono
        self.correo = correo
        self.fecha_inicio = fecha_inicio
        self.salario = salario
        self.id_tipo = id_tipo
        self.id_departamento = id_departamento
        self.id_rol = id_rol

    def obtener_info_empleado():
        # Solicitar ID de empleado y verificar que no exista
        while True:
            id_empleado = input("Ingrese el ID del empleado: ")

            with mydb.cursor() as cursor:
                cursor.execute("SELECT id_empleado FROM empleado WHERE id_empleado = %s", (id_empleado,))
                if cursor.fetchone():
                    print("❌ El ID de empleado ya existe. Intente con otro.")
                else:
                    break  # Si no existe, continuar solicitando la información del empleado

        nombre = input("Ingrese el nombre del empleado: ")
        direccion = input("Ingrese la dirección del empleado: ")
        telefono = input("Ingrese el número de teléfono del empleado: ")
        correo = input("Ingrese el correo del empleado: ")
        fecha_inicio = input("Ingrese la fecha de inicio (YYYY-MM-DD): ")

        # Validación del salario
        while True:
            try:
                salario = float(input("Ingrese el salario del empleado: "))
                if salario < 0:
                    raise ValueError("El salario no puede ser negativo.")
                break
            except ValueError as e:
                print(f"Entrada inválida: {e}. Intente nuevamente.")

        id_tipo = input("Ingrese el ID del tipo de empleado: ")

        # Verificación y entrada del ID de departamento
        while True:
            id_departamento_input = input("Ingrese el ID del departamento: ")
            with mydb.cursor() as cursor:
                cursor.execute("SELECT habilitado FROM departamentos WHERE id_departamento = %s", (id_departamento_input,))
                resultado = cursor.fetchone()
                if resultado:
                    if resultado[0] == 1:  # Verifica si está habilitado
                        id_departamento = id_departamento_input
                        break
                    else:
                        print("❌ El departamento está deshabilitado. Intente con otro ID.")
                else:
                    print("❌ El ID del departamento no existe. Intente nuevamente.")

        # Verificación y entrada del ID de rol
        while True:
            id_rol_input = input("Ingrese el ID del rol: ")
            with mydb.cursor() as cursor:
                cursor.execute("SELECT id_rol, habilitado FROM roles WHERE id_rol = %s", (id_rol_input,))
                resultado = cursor.fetchone()
                if resultado:
                    if resultado[1] == 1:  # Verifica si el rol está habilitado
                        id_rol = id_rol_input
                        break
                    else:
                        print("❌ El rol está deshabilitado. Intente con otro ID.")
                else:
                    print("❌ El ID de rol no existe. Intente nuevamente.")

        try:
            with mydb.cursor() as miCursor:
                sql = """
                    INSERT INTO empleado (id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento, id_rol)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                val = (id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento, id_rol)
                miCursor.execute(sql, val)
                mydb.commit()
                print(miCursor.rowcount, "Registro exitoso")
                # Devolver un objeto de tipo Empleado
                return Empleado(id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento, id_rol)
        except Exception as e:
            print(f"Ocurrió un error al registrar el empleado: {e}")
            return None
        
    def mostrar_empleados(): #solo muestra los empleados habilitaods
        miCusor = mydb.cursor()
        sql = "SELECT id_empleado, nombre, direccion, telefono, correo, fecha_inicio, salario, id_tipo, id_departamento FROM empleado WHERE habilitado = 1;"
        miCusor.execute(sql)
        empleados = miCusor.fetchall()

        # Crear la tabla
        tabla = PrettyTable()
        print("Empleados:")
        tabla.field_names = ["ID", "Nombre", "Dirección", "Teléfono", "Correo", "Fecha Inicio", "Salario", "ID Tipo", "ID Departamento"]

        # Añadir filas a la tabla
        for empleado in empleados:
            tabla.add_row(empleado)

        print(tabla)


    def buscar_empleado(id_empleado):
        try:
            miCursor = mydb.cursor()
            sql = "SELECT * FROM empleado WHERE id_empleado = %s"
            miCursor.execute(sql, (id_empleado,))
            empleado = miCursor.fetchone()

            if empleado:
                habilitado = "Sí" if empleado[9] else "No"  # Asumiendo que 'habilitado' es el décimo atributo
                print(f"Empleado encontrado: ID: {empleado[0]}, Nombre: {empleado[1]}, Dirección: {empleado[2]}, Teléfono: {empleado[3]}, Correo: {empleado[4]}, Fecha de Inicio: {empleado[5]}, Salario: {empleado[6]}, ID Tipo: {empleado[7]}, ID Departamento: {empleado[8]}, Habilitado: {habilitado}")
            else:
                print("Empleado no encontrado.")
        except Exception as e:
            print(f"Ocurrió un error: {e}")
        finally:
            miCursor.close()
        
    def editar_empleado(id_empleado):
        miCusor = mydb.cursor()
        sql = "SELECT * FROM empleado WHERE id_empleado = %s"
        miCusor.execute(sql, (id_empleado,))
        empleado = miCusor.fetchone()

        if empleado:
            print("Empleado encontrado:")
            print(f"ID: {empleado[0]}, Nombre: {empleado[1]}, Dirección: {empleado[2]}, Teléfono: {empleado[3]}, Correo: {empleado[4]}, Salario: {empleado[6]}")

            # Preguntar qué desea cambiar
            nuevo_direccion = input("Nueva dirección (dejar vacío para no cambiar): ")
            nuevo_correo = input("Nuevo correo (dejar vacío para no cambiar): ")

            # Validar y preguntar el número de teléfono hasta que sea correcto
            while True:
                nuevo_telefono = input("Nuevo teléfono (dejar vacío para no cambiar): ")
                if not nuevo_telefono:  # Teléfono vacío
                    break
                try:
                    # Validar que el teléfono sea un número y convertirlo a int
                    int(nuevo_telefono)
                    break  # Teléfono válido, salir del bucle
                except ValueError:
                    print("Error: El teléfono debe ser un número válido. Intente nuevamente.")

            # Validar y preguntar el salario hasta que sea correcto
            while True:
                nuevo_salario = input("Nuevo salario (dejar vacío para no cambiar): ")
                if not nuevo_salario:  # Salario vacío
                    break
                try:
                    nuevo_salario_float = float(nuevo_salario)
                    break  # Salario válido, salir del bucle
                except ValueError:
                    print("Error: El salario debe ser un número válido. Intente nuevamente.")

            # Actualizar solo los campos que el usuario ha proporcionado
            if nuevo_direccion:
                sql_update = "UPDATE empleado SET direccion = %s WHERE id_empleado = %s"
                miCusor.execute(sql_update, (nuevo_direccion, id_empleado))
                
            if nuevo_correo:
                sql_update = "UPDATE empleado SET correo = %s WHERE id_empleado = %s"
                miCusor.execute(sql_update, (nuevo_correo, id_empleado))

            if nuevo_telefono:  # Solo se actualiza si se proporcionó un nuevo teléfono
                sql_update = "UPDATE empleado SET telefono = %s WHERE id_empleado = %s"
                miCusor.execute(sql_update, (nuevo_telefono, id_empleado))

            if nuevo_salario:  # Solo se actualiza si se proporcionó un nuevo salario
                sql_update = "UPDATE empleado SET salario = %s WHERE id_empleado = %s"
                miCusor.execute(sql_update, (nuevo_salario_float, id_empleado))

            mydb.commit()
            print("Empleado actualizado exitosamente.")
        else:
            print("Empleado no encontrado.")


    def eliminar_empleado(id_empleado):
        miCursor = mydb.cursor()
        # Verificar si el empleado existe
        sql_select = "SELECT habilitado FROM empleado WHERE id_empleado = %s"
        miCursor.execute(sql_select, (id_empleado,))
        resultado = miCursor.fetchone()
        
        if resultado is not None and resultado[0]:  # Si el empleado existe y está habilitado
            desabilitado = 0
            sql_update = "UPDATE empleado SET habilitado = %s WHERE id_empleado = %s"
            miCursor.execute(sql_update, (desabilitado, id_empleado))
            mydb.commit()
            print("Empleado deshabilitado correctamente.")
        else:
            print("Empleado no encontrado o ya ha sido deshabilitado.")

        miCursor.close()



    def informe_empleados():
        try:
            with mydb.cursor() as miCursor:
                sql = "SELECT * FROM empleado"
                miCursor.execute(sql)
                resultados = miCursor.fetchall()

                # Crear un DataFrame para organizar los datos
                datos = []
                for empleado in resultados:
                    datos.append({
                        "ID": empleado[0],
                        "Nombre": empleado[1],
                        "Dirección": empleado[2],
                        "Teléfono": empleado[3],
                        "Correo": empleado[4],
                        "Fecha de Inicio": empleado[5],
                        "Salario": empleado[6],
                        "ID Tipo": empleado[7],
                        "ID Departamento": empleado[8],
                    })

                df = pd.DataFrame(datos)

                # Crear un nuevo archivo Excel
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Empleados"

                # Escribir encabezados
                for col in range(len(df.columns)):
                    cell = sheet.cell(row=1, column=col + 1)
                    cell.value = df.columns[col]
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Escribir datos
                for row in range(len(df)):
                    for col in range(len(df.columns)):
                        cell = sheet.cell(row=row + 2, column=col + 1)
                        cell.value = df.iat[row, col]
                        cell.alignment = Alignment(horizontal='left')
                    
                # Ajustar el ancho de las columnas
                for column in sheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column[0].column_letter].width = adjusted_width

                # Guardar el archivo
                workbook.save("informe_empleados.xlsx")

                print("Informe de empleados generado exitosamente en 'informe_empleados.xlsx'.")

        except Exception as e:
            print(f"Ocurrió un error al generar el informe: {e}")

    def reasignar_departamento(id_empleado, nuevo_id_departamento):
        miCursor = mydb.cursor()
        # Verificar si el empleado existe
        sql = "SELECT * FROM empleado WHERE id_empleado = %s"
        miCursor.execute(sql, (id_empleado,))
        empleado = miCursor.fetchone()

        if empleado:
            # Actualizar el ID del departamento
            sql_update = "UPDATE empleado SET id_departamento = %s WHERE id_empleado = %s"
            miCursor.execute(sql_update, (nuevo_id_departamento, id_empleado))
            mydb.commit()
            print(f"El empleado con ID {id_empleado} ha sido reasignado al departamento {nuevo_id_departamento}.")
        else:
            print("Empleado no encontrado.")

    def mostrar_resumen_empleados():
        miCursor = mydb.cursor()
        sql = """
            SELECT e.id_empleado, e.nombre, e.id_departamento, d.nombre_departamentos
            FROM empleado e
            JOIN departamentos d ON e.id_departamento = d.id_departamento
        """
        miCursor.execute(sql)
        empleados = miCursor.fetchall()

        # Crear la tabla
        tabla = PrettyTable()
        print("Resumen de Empleados:")
        tabla.field_names = ["ID", "Nombre", "ID Departamento", "Nombre Departamento"]

        # Añadir filas a la tabla
        for empleado in empleados:
            tabla.add_row([empleado[0], empleado[1], empleado[2], empleado[3]])

        print(tabla)
